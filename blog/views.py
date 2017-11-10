"""views.py."""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate
from django.forms.formsets import formset_factory
from django.utils import timezone
from django.db import IntegrityError, transaction
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import smtplib
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from .models import Inventory, Order, Product, Computing, Electronic, Optic, Chemical, Biological
from .models import Instrumentation, Others, Full_Name_Users, Run, Chip, Wafer, Waveguide
from .models import Name_Waveguide, Type_Component, Type_Optic, Type_Chemical, State
from .models import Type_Biological_1, Type_Biological_2, Type_Instrumentation
from .forms import InventoryForm, OrderForm, ProductForm, ComputingForm
from .forms import ElectronicForm, OpticForm, ChemicalForm, BiologicalForm, InstrumentationForm
from .forms import OthersForm, SignUpForm, RunForm, WaferForm, ChipForm, WaveguideForm
from .forms import SendEmailForm


def home(request):
    """
    Home function docstring.

    This function shows the main page.

    @param request: HTML request page.

    @return: Main page.
    """
    return render(request, 'blog/home.html')


def signup(request):
    """
    Signup function docstring.

    This function carries out the registration of a new user.

    @param request: HTML request page.

    @return: First time, return to sign up page. If the sign up form is completed, return to home
    page with your username.
    """
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            firstname = form.cleaned_data.get('first_name')
            lastname = form.cleaned_data.get('last_name')
            nameFull = firstname + ' ' + lastname
            fullName = Full_Name_Users()
            fullName.name = nameFull
            fullName.save()
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            return redirect('blog:home')
    else:
        form = SignUpForm()
    return render(request, 'registration/signup.html', {'form': form})


def inventory_list(request):
    """
    Inventory_list function docstring.

    This function shows the different inventories of this web app that are ordered by creation date.

    @param request: HTML request page.

    @return: list of inventories.
    """
    inventories = Inventory.objects.filter(
        created_date__lte=timezone.now()).order_by('created_date').reverse()

    return render(request, 'blog/inventory_list.html', {'inventories': inventories})


def inventory_detail(request, pk):
    """
    Inventory_detail function docstring.

    This function shows the information of an inventory.

    @param request: HTML request page.

    @param pk: primary key of the inventory.

    @return: one inventory.

    @raise 404: inventory does not exists.
    """
    inventory = get_object_or_404(Inventory, pk=pk)

    return render(request, 'blog/inventory_detail.html', {'inventory': inventory})


@login_required
def inventory_new(request):
    """
    Inventory_new function docstring.

    This function shows the form to create a new inventory.

    @param request: HTML request page.

    @return: First time, this shows the form to a new inventory. If the form is completed, return
    the details of this new inventory.
    """
    if request.method == "POST":
        form = InventoryForm(request.POST)
        if form.is_valid():
            inventory = form.save(commit=False)
            inventory.author = request.user
            # inventory.published_date = timezone.now()
            inventory.save()
            return redirect('blog:inventory_detail', pk=inventory.pk)
    else:
        form = InventoryForm()
    return render(request, 'blog/inventory_edit.html', {'form': form})


@login_required
def inventory_edit(request, pk):
    """
    Inventory_edit function docstring.

    This function shows the form to modify an inventory.

    @param request: HTML request page.

    @param pk: primary key of the inventory to modify.

    @return: First time, this shows the form to edit the inventory. If the form is completed, return
    the details of this inventory.

    @raise 404: inventory does not exists.
    """
    inventory = get_object_or_404(Inventory, pk=pk)
    if request.method == "POST":
        form = InventoryForm(data=request.POST, instance=inventory)
        if form.is_valid():
            inventory = form.save(commit=False)
            inventory.author = request.user
            inventory.save()
            return redirect('blog:inventory_detail', pk=inventory.pk)
    else:
        form = InventoryForm(instance=inventory)
    return render(request, 'blog/inventory_edit.html', {'form': form})


@login_required
def inventory_remove(request, pk):
    """
    Inventory_remove function docstring.

    This function removes an inventory.

    @param request: HTML request page.

    @param pk: primary key of the inventory to remove.

    @return: list of inventories.

    @raise 404: inventory does not exists.
    """
    inventory = get_object_or_404(Inventory, pk=pk)
    inventory.delete()
    return redirect('blog:inventory_list')


def computing_list(request):
    """
    Computing_list function docstring.

    This function shows the list of computers that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.

    @return: list of computers.
    """
    computings = Computing.objects.filter(
        created_date__lte=timezone.now()).order_by('created_date').reverse()

    return render(request, 'blog/computing_list.html', {'computings': computings})


def computing_detail(request, pk):
    """
    Computing_detail function docstring.

    This function shows the information of a computer.

    @param request: HTML request page.

    @param pk: primary key of the computer.

    @return: one computer.

    @raise 404: computer does not exists.
    """
    computing = get_object_or_404(Computing, pk=pk)

    return render(request, 'blog/computing_detail.html', {'computing': computing})


@login_required
def computing_new(request):
    """
    Computing_new function docstring.

    This function shows the form to create a new computer.

    @param request: HTML request page.

    @return: First time, this shows the form to a new computer. If the form is completed, return the
    details of this new computer.
    """
    if request.method == "POST":
        form = ComputingForm(request.POST)
        if form.is_valid():
            computing = form.save(commit=False)
            computing.save()
            return redirect('blog:computing_detail', pk=computing.pk)
    else:
        form = ComputingForm()
    return render(request, 'blog/computing_new.html', {'form': form})


@login_required
def computing_edit(request, pk):
    """
    Computing_edit function docstring.

    This function shows the form to modify a computer.

    @param request: HTML request page.

    @param pk: primary key of the computer to modify.

    @return: First time, this shows the form to edit the computer information. If the form is
    completed, return the details of this computer.

    @raise 404: computer does not exists.
    """
    computing = get_object_or_404(Computing, pk=pk)
    if request.method == "POST":
        computing_form = ComputingForm(data=request.POST, instance=computing)
        if computing_form.is_valid():
            computing = computing_form.save(commit=False)
            computing_all = Computing.objects.all()

            duplicates = False

            for data in computing_all:
                if data.name == computing.name and data.pk != computing.pk:
                    duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your computing.')
                computing.save()
                return redirect('blog:computing_detail', pk=computing.pk)
            else:
                messages.warning(request, 'Already exists a computing with this name.')
                return redirect('blog:computing_edit', pk=computing.pk)

    else:
        form = ComputingForm(instance=computing)
    return render(request, 'blog/computing_edit.html', {'form': form})


@login_required
def computing_remove(request, pk):
    """
    Computing_remove function docstring.

    This function removes a computer.

    @param request: HTML request page.

    @param pk: primary key of the computer to remove.

    @return: list of computers.

    @raise 404: computer does not exists.
    """
    computing = get_object_or_404(Computing, pk=pk)
    computing.delete()
    return redirect('blog:computing_list')


def electronic_list_type_components(request):
    """
    electronic_list_type_components function docstring.

    This function shows the list of type components that are stored in this web app and
    they are ordered by creation date.

    @param request: HTML request page.

    @return: list of type components.
    """
    # electronics = Electronic.objects.filter(
    #     created_date__lte=timezone.now()).order_by('created_date').reverse()

    electronics = Type_Component.objects.all()

    return render(request, 'blog/electronic_list_type_components.html',
                  {'electronics': electronics})


def electronic_list(request, pk):
    """
    Electronic_list function docstring.

    This function shows the list of electronic components that are stored in this web app and
    they are ordered by creation date.

    @param request: HTML request page.
    @param pk: primary key of chemical type.

    @return: list of electronic components.
    """
    type_component = Type_Component.objects.get(pk=pk)
    electronics = Electronic.objects.filter(type_component=type_component)
    componentsBack = True
    type_componBack = False

    return render(request, 'blog/electronic_list.html', {'electronics': electronics,
                                                         'componentsBack': componentsBack,
                                                         'type_componBack': type_componBack})


def electronic_detail(request, pk):
    """
    Electronic_detail function docstring.

    This function shows the information of a electronic component.

    @param request: HTML request page.

    @param pk: primary key of the electronic component.

    @return: one electronic component.

    @raise 404: electronic component does not exists.
    """
    electronic = get_object_or_404(Electronic, pk=pk)
    componentsBack = False
    type_componBack = True

    return render(request, 'blog/electronic_detail.html', {'electronic': electronic,
                                                           'componentsBack': componentsBack,
                                                           'type_componBack': type_componBack})


@login_required
def electronic_new(request):
    """
    Electronic_new function docstring.

    This function shows the form to create a new electronic component.

    @param request: HTML request page.

    @return: First time, this shows the form to a new electronic component. If the form is
    completed, return the details of this new electronic component.
    """
    if request.method == "POST":
        form = ElectronicForm(request.POST)
        if form.is_valid():
            electronic = form.save(commit=False)
            electronic.name_component = electronic.type_component.name
            electronic.save()
            return redirect('blog:electronic_detail', pk=electronic.pk)
    else:
        form = ElectronicForm()
    return render(request, 'blog/electronic_new.html', {'form': form})


@login_required
def electronic_edit(request, pk):
    """
    Electronic_edit function docstring.

    This function shows the form to modify a electronic component.

    @param request: HTML request page.

    @param pk: primary key of the electronic component to modify.

    @return: First time, this shows the form to edit the electronic component information. If the
    form is completed, return the details of this electronic component.

    @raise 404: electronic component does not exists.
    """
    electronic = get_object_or_404(Electronic, pk=pk)
    if request.method == "POST":
        form = ElectronicForm(data=request.POST, instance=electronic)
        if form.is_valid():
            electronic = form.save(commit=False)
            electronic_all = Electronic.objects.all()

            duplicates = False

            for data in electronic_all:
                if data.value == electronic.value and data.unit == electronic.unit:
                    if data.pk != electronic.pk:
                        duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your electronic component.')
                electronic.save()
                return redirect('blog:electronic_detail', pk=electronic.pk)
            else:
                messages.warning(request, 'Already exists an electronic component with this name.')
                return redirect('blog:electronic_edit', pk=electronic.pk)

    else:
        form = ElectronicForm(instance=electronic)
    return render(request, 'blog/electronic_edit.html', {'form': form})


@login_required
def electronic_remove(request, pk):
    """
    Electronic_remove function docstring.

    This function removes a electronic component.

    @param request: HTML request page.

    @param pk: primary key of the electronic component to remove.

    @return: list of electronic components.

    @raise 404: electronic component does not exists.
    """
    electronic = get_object_or_404(Electronic, pk=pk)
    name_component = electronic.name_component
    electronic.delete()
    return redirect('blog:electronic_list', component=name_component)


def optic_list_type_optic(request):
    """
    Optic_list_type_optic function docstring.

    This function shows the list of type of optic components that are stored in this web app and
    they are ordered by creation date.

    @param request: HTML request page.

    @return: list of type of optic components.
    """
    optics = Type_Optic.objects.all()

    return render(request, 'blog/optic_list_type_optic.html', {'optics': optics})


def optic_list(request, pk):
    """
    Optic_list function docstring.

    This function shows the list of optic components that are stored in this web app and
    they are ordered by creation date.

    @param request: HTML request page.
    @param pk: primary key of chemical type.

    @return: list of optic components.
    """
    type_optic = Type_Optic.objects.get(pk=pk)
    optics = Optic.objects.filter(type_optic=type_optic)
    opticsBack = True
    type_opticBack = False

    return render(request, 'blog/optic_list.html', {'optics': optics, 'opticsBack': opticsBack,
                                                    'type_opticBack': type_opticBack})


def optic_detail(request, pk):
    """
    Optic_detail function docstring.

    This function shows the information of a optic component.

    @param request: HTML request page.

    @param pk: primary key of the optic component.

    @return: one optic component.

    @raise 404: optic component does not exists.
    """
    optic = get_object_or_404(Optic, pk=pk)
    opticsBack = False
    type_opticBack = True

    return render(request, 'blog/optic_detail.html', {'optic': optic, 'opticsBack': opticsBack,
                                                      'type_opticBack': type_opticBack})


@login_required
def optic_new(request):
    """
    Optic_new function docstring.

    This function shows the form to create a new optic component.

    @param request: HTML request page.

    @return: First time, this shows the form to a new optic component. If the form is
    completed, return the details of this new optic component.
    """
    if request.method == "POST":
        form = OpticForm(request.POST)
        if form.is_valid():
            optic = form.save(commit=False)
            optic.name_optic = optic.type_optic.name
            optic.save()
            return redirect('blog:optic_detail', pk=optic.pk)
    else:
        form = OpticForm()
    return render(request, 'blog/optic_new.html', {'form': form})


@login_required
def optic_edit(request, pk):
    """
    Optic_edit function docstring.

    This function shows the form to modify a optic component.

    @param request: HTML request page.

    @param pk: primary key of the optic component to modify.

    @return: First time, this shows the form to edit the optic component information. If the
    form is completed, return the details of this optic component.

    @raise 404: optic component does not exists.
    """
    optic = get_object_or_404(Optic, pk=pk)
    if request.method == "POST":
        form = OpticForm(data=request.POST, instance=optic)
        if form.is_valid():
            optic = form.save(commit=False)
            optic_all = Optic.objects.all()

            duplicates = False

            for data in optic_all:
                if data.description == optic.description and data.pk != optic.pk:
                    duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your optic component.')
                optic.save()
                return redirect('blog:optic_detail', pk=optic.pk)
            else:
                messages.warning(request, 'Already exists an optic component with this name.')
                return redirect('blog:optic_edit', pk=optic.pk)
    else:
        form = OpticForm(instance=optic)
    return render(request, 'blog/optic_edit.html', {'form': form})


@login_required
def optic_remove(request, pk):
    """
    Optic_remove function docstring.

    This function removes a optic component.

    @param request: HTML request page.

    @param pk: primary key of the optic component to remove.

    @return: list of optic components.

    @raise 404: optic component does not exists.
    """
    optic = get_object_or_404(Optic, pk=pk)
    optic.delete()
    return redirect('blog:optic_list')


def chemical_list_type_chemical(request):
    """
    Chemical_list_type_chemical function docstring.

    This function shows the list of type of chemicals that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.

    @return: list of type of chemicals.
    """
    states = State.objects.all()
    chemicals = Type_Chemical.objects.all()

    return render(request, 'blog/chemical_list_type_chemical.html', {'states': states,
                                                                     'chemicals': chemicals})


def chemical_list(request, pk):
    """
    Chemical_list function docstring.

    This function shows the list of chemicals that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.
    @param pk: primary key of chemical type.

    @return: list of chemicals.
    """
    type_chemical = Type_Chemical.objects.get(pk=pk)
    chemicals = Chemical.objects.filter(type_chemical=type_chemical)
    chemicalsBack = True
    type_chemicalBack = False

    return render(request, 'blog/chemical_list.html', {'chemicals': chemicals,
                                                       'chemicalsBack': chemicalsBack,
                                                       'type_chemicalBack': type_chemicalBack})


def chemical_detail(request, pk):
    """
    Chemical_detail function docstring.

    This function shows the information of a chemical.

    @param request: HTML request page.

    @param pk: primary key of the chemical.

    @return: one chemical.

    @raise 404: chemical does not exists.
    """
    chemical = get_object_or_404(Chemical, pk=pk)
    chemicalsBack = False
    type_chemicalBack = True

    return render(request, 'blog/chemical_detail.html', {'chemical': chemical,
                                                         'chemicalsBack': chemicalsBack,
                                                         'type_chemicalBack': type_chemicalBack})


@login_required
def chemical_new(request):
    """
    Chemical_new function docstring.

    This function shows the form to create a new chemical.

    @param request: HTML request page.

    @return: First time, this shows the form to a new chemical. If the form is completed, return
    the details of this new chemical.
    """
    if request.method == "POST":
        form = ChemicalForm(request.POST)
        if form.is_valid():
            chemical = form.save(commit=False)
            chemical.save()
            return redirect('blog:chemical_detail', pk=chemical.pk)
    else:
        form = ChemicalForm()
    return render(request, 'blog/chemical_new.html', {'form': form})


@login_required
def chemical_edit(request, pk):
    """
    Chemical_edit function docstring.

    This function shows the form to modify a chemical.

    @param request: HTML request page.

    @param pk: primary key of the chemical to modify.

    @return: First time, this shows the form to edit the chemical information. If the form is
    completed, return the details of this chemical.

    @raise 404: chemical does not exists.
    """
    chemical = get_object_or_404(Chemical, pk=pk)
    if request.method == "POST":
        form = ChemicalForm(data=request.POST, instance=chemical)
        if form.is_valid():
            chemical = form.save(commit=False)
            chemical_all = Chemical.objects.all()

            duplicates = False

            for data in chemical_all:
                if data.name == chemical.name and data.pk != chemical.pk:
                    duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your chemical.')
                chemical.save()
                return redirect('blog:chemical_detail', pk=chemical.pk)
            else:
                messages.warning(request, 'Already exists an chemical with this name.')
                return redirect('blog:chemical_edit', pk=chemical.pk)

    else:
        form = ChemicalForm(instance=chemical)
    return render(request, 'blog/chemical_edit.html', {'form': form})


@login_required
def chemical_remove(request, pk):
    """
    Chemical_remove function docstring.

    This function removes a chemical.

    @param request: HTML request page.

    @param pk: primary key of the chemical to remove.

    @return: list of chemicals.

    @raise 404: chemical does not exists.
    """
    chemical = get_object_or_404(Chemical, pk=pk)
    chemical.delete()
    return redirect('blog:chemical_list')


def biological_list_type_biological(request):
    """
    biological_list_type_biological function docstring.

    This function shows the list of type of biologicals that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.

    @return: list of type of biologicals.
    """
    biologicals = Type_Biological_1.objects.all()

    biologicals_2 = []

    for biological in biologicals:
        data = Type_Biological_2.objects.filter(type_biological_1=biological)
        biologicals_2.append(data)

    # for bios in biologicals_2:
    #     for bio in bios:
    #         print(bio.type_biological_1)

    return render(request, 'blog/biological_list_type_biological.html',
                  {'biologicals': biologicals, 'biologicals_2': biologicals_2})


def biological_list(request, pk):
    """
    Biological_list function docstring.

    This function shows the list of biologicals that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.
    @param pk: primary key of biological types.

    @return: list of biological types.
    """
    type_biological_2 = Type_Biological_2.objects.get(pk=pk)
    biologicals = Biological.objects.filter(type_biological=type_biological_2)
    print(biologicals)
    biologicalsBack = True
    type_bioBack = False

    return render(request, 'blog/biological_list.html', {'biologicals': biologicals,
                                                         'biologicalsBack': biologicalsBack,
                                                         'type_bioBack': type_bioBack})


def biological_detail(request, pk):
    """
    Biological_detail function docstring.

    This function shows the information of a biological.

    @param request: HTML request page.

    @param pk: primary key of the biological.

    @return: one biological.

    @raise 404: biological does not exists.
    """
    biological = get_object_or_404(Biological, pk=pk)
    biologicalsBack = False
    type_bioBack = True

    return render(request, 'blog/biological_detail.html', {'biological': biological,
                                                           'biologicalsBack': biologicalsBack,
                                                           'type_bioBack': type_bioBack})


@login_required
def biological_new(request):
    """
    Biological_new function docstring.

    This function shows the form to create a new biological.

    @param request: HTML request page.

    @return: First time, this shows the form to a new biological. If the form is completed, return
    the details of this new biological.
    """
    if request.method == "POST":
        form = BiologicalForm(request.POST)
        if form.is_valid():
            biological = form.save(commit=False)
            biological.save()
            return redirect('blog:biological_detail', pk=biological.pk)
    else:
        form = BiologicalForm()
    return render(request, 'blog/biological_new.html', {'form': form})


@login_required
def biological_edit(request, pk):
    """
    Biological_edit function docstring.

    This function shows the form to modify a biological.

    @param request: HTML request page.

    @param pk: primary key of the biological to modify.

    @return: First time, this shows the form to edit the biological information. If the form is
    completed, return the details of this biological.

    @raise 404: biological does not exists.
    """
    biological = get_object_or_404(Biological, pk=pk)
    if request.method == "POST":
        form = BiologicalForm(data=request.POST, instance=biological)
        if form.is_valid():
            biological = form.save(commit=False)
            biological_all = Biological.objects.all()

            duplicates = False

            for data in biological_all:
                if data.name == biological.name and data.pk != biological.pk:
                    duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your biological.')
                biological.save()
                return redirect('blog:biological_detail', pk=biological.pk)
            else:
                messages.warning(request, 'Already exists an biological with this name.')
                return redirect('blog:biological_edit', pk=biological.pk)

    else:
        form = BiologicalForm(instance=biological)
    return render(request, 'blog/biological_edit.html', {'form': form})


@login_required
def biological_remove(request, pk):
    """
    Biological_remove function docstring.

    This function removes a biological.

    @param request: HTML request page.

    @param pk: primary key of the biological to remove.

    @return: list of biologicals.

    @raise 404: biological does not exists.
    """
    biological = get_object_or_404(Biological, pk=pk)
    biological.delete()
    return redirect('blog:biological_list')


def instrumentation_list_type(request):
    """
    Instrumentation_list_type function docstring.

    This function shows the list of instrument types that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.

    @return: list of instrument types.
    """
    instrumentations = Type_Instrumentation.objects.all()

    return render(request, 'blog/instrumentation_list_type.html',
                  {'instrumentations': instrumentations})


def instrumentation_list(request, pk):
    """
    Instrumentation_list function docstring.

    This function shows the list of instruments that are stored in this web app and they are
    ordered by creation date.

    @param request: HTML request page.
    @param pk: primary key of instrument types..

    @return: list of instrument types.
    """
    type_instrumentation = Type_Instrumentation.objects.get(pk=pk)
    instrumentations = Instrumentation.objects.filter(type_instrumentation=type_instrumentation)
    instrumentBack = True
    type_instrumBack = False

    return render(request, 'blog/instrumentation_list.html', {'instrumentations': instrumentations,
                                                              'instrumentBack': instrumentBack,
                                                              'type_instrumBack': type_instrumBack})


def instrumentation_detail(request, pk):
    """
    Instrumentation_detail function docstring.

    This function shows the information of a instrument.

    @param request: HTML request page.

    @param pk: primary key of the instrument.

    @return: one instrument.

    @raise 404: instrument does not exists.
    """
    instrumentation = get_object_or_404(Instrumentation, pk=pk)
    instrumentBack = False
    type_instrumBack = True

    return render(request, 'blog/instrumentation_detail.html', {'instrumentation': instrumentation,
                                                                'instrumentBack': instrumentBack,
                                                                'type_instrumBack': type_instrumBack
                                                                })


@login_required
def instrumentation_new(request):
    """
    Instrumentation_new function docstring.

    This function shows the form to create a new instrument.

    @param request: HTML request page.

    @return: First time, this shows the form to a new instrument. If the form is completed, return
    the details of this new instrument.
    """
    if request.method == "POST":
        form = InstrumentationForm(request.POST)
        if form.is_valid():
            instrumentation = form.save(commit=False)
            instrumentation.save()
            return redirect('blog:instrumentation_detail', pk=instrumentation.pk)
    else:
        form = InstrumentationForm()
    return render(request, 'blog/instrumentation_new.html', {'form': form})


@login_required
def instrumentation_edit(request, pk):
    """
    Instrumentation_edit function docstring.

    This function shows the form to modify a instrument.

    @param request: HTML request page.

    @param pk: primary key of the instrument to modify.

    @return: First time, this shows the form to edit the instrument information. If the form is
    completed, return the details of this instrument.

    @raise 404: instrument does not exists.
    """
    instrumentation = get_object_or_404(Instrumentation, pk=pk)
    if request.method == "POST":
        form = InstrumentationForm(data=request.POST, instance=instrumentation)
        if form.is_valid():
            instrumentation = form.save(commit=False)
            instrumentation_all = Instrumentation.objects.all()

            duplicates = False

            for data in instrumentation_all:
                if data.characteristics == instrumentation.characteristics:
                    if data.pk != instrumentation.pk:
                        duplicates = True

            if not duplicates:
                messages.success(request, 'You have updated your instrumentation.')
                instrumentation.save()
                return redirect('blog:instrumentation_detail', pk=instrumentation.pk)
            else:
                messages.warning(request, 'Already exists an instrumentation with this name.')
                return redirect('blog:instrumentation_edit', pk=instrumentation.pk)

    else:
        form = InstrumentationForm(instance=instrumentation)
    return render(request, 'blog/instrumentation_edit.html', {'form': form})


@login_required
def instrumentation_remove(request, pk):
    """
    Instrumentation_remove function docstring.

    This function removes a instrument.

    @param request: HTML request page.

    @param pk: primary key of the instrument to remove.

    @return: list of instruments.

    @raise 404: instrument does not exists.
    """
    instrumentation = get_object_or_404(Instrumentation, pk=pk)
    instrumentation.delete()
    return redirect('blog:instrumentation_list')


def others_list(request):
    """
    Others_list function docstring.

    This function shows the list of components whithout type that are stored in this web app and
    they are ordered by creation date.

    @param request: HTML request page.

    @return: list of components whithout type.
    """
    otherss = Others.objects.filter(
        created_date__lte=timezone.now()).order_by('created_date').reverse()

    return render(request, 'blog/others_list.html', {'otherss': otherss})


def others_detail(request, pk):
    """
    Others_detail function docstring.

    This function shows the information of a component whithout type.

    @param request: HTML request page.

    @param pk: primary key of the component.

    @return: one components whithout type.

    @raise 404: component does not exists.
    """
    others = get_object_or_404(Others, pk=pk)

    return render(request, 'blog/others_detail.html', {'others': others})


@login_required
def others_new(request):
    """
    Others_new function docstring.

    This function shows the form to create a new component without type.

    @param request: HTML request page.

    @return: First time, this shows the form to a new component. If the form is completed, return
    the details of this new component without type.
    """
    if request.method == "POST":
        form = OthersForm(request.POST)
        if form.is_valid():
            others = form.save(commit=False)
            others.save()
            return redirect('blog:others_detail', pk=others.pk)
    else:
        form = OthersForm()
    return render(request, 'blog/others_edit.html', {'form': form})


@login_required
def others_edit(request, pk):
    """
    Others_edit function docstring.

    This function shows the form to modify a component without type.

    @param request: HTML request page.

    @param pk: primary key of the component to modify.

    @return: First time, this shows the form to edit the component information. If the form is
    completed, return the details of this component without type.

    @raise 404: component does not exists.
    """
    others = get_object_or_404(Others, pk=pk)
    if request.method == "POST":
        form = OthersForm(data=request.POST, instance=others)
        if form.is_valid():
            others = form.save(commit=False)
            others.save()
            return redirect('blog:others_detail', pk=others.pk)
    else:
        form = OthersForm(instance=others)
    return render(request, 'blog/others_edit.html', {'form': form})


@login_required
def others_remove(request, pk):
    """
    Others_remove function docstring.

    This function removes a component that does not have type.

    @param request: HTML request page.

    @param pk: primary key of the component to remove.

    @return: list of components.

    @raise 404: component does not exists.
    """
    others = get_object_or_404(Others, pk=pk)
    others.delete()
    return redirect('blog:others_list')


def order_list(request):
    """
    Order_list function docstring.

    This function shows the list of orders carried out in this web app and they are ordered by
    creation date.

    @param request: HTML request page.

    @return: list of orders.
    """
    orders = Order.objects.filter(
        created_date__lte=timezone.now()).order_by('created_date').reverse()

    return render(request, 'blog/order_list.html', {'orders': orders})


def order_detail(request, pk):
    """
    Order_detail function docstring.

    This function shows the information of an order.

    @param request: HTML request page.

    @param pk: primary key of the order.

    @return: one order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk).order_by('created_date')

    noItem = False
    if not products.exists():
        noItem = True

    return render(request, 'blog/order_detail.html', {'order': order, 'products': products,
                                                      'noItem': noItem})


@login_required
def order_new(request):
    """
    Order_new function docstring.

    This function shows the first step of the form to create a new order.

    @param request: HTML request page.

    @return: First time, this shows the form to a new order. If the form is completed, return
    the next step to finish the order.
    """
    ProductFormSet = formset_factory(ProductForm)

    if request.method == "POST":
        order_form = OrderForm(data=request.POST, prefix="orderForm")
        product_formset = ProductFormSet(data=request.POST, prefix="form")

        print(order_form)
        print(product_formset)

        if order_form.is_valid() and product_formset.is_valid():
            order = order_form.save(commit=False)
            order.author = request.user
            order.save()

            new_products = []
            duplicates = False

            for product_form in product_formset:
                description = product_form.cleaned_data.get('description')
                quantity = product_form.cleaned_data.get('quantity')
                unit_price = product_form.cleaned_data.get('unit_price')

                if description and quantity and unit_price:
                    for new_products_data in new_products:
                        if new_products_data.description == description:
                            duplicates = True

                    new_products.append(Product(description=description, quantity=quantity,
                                                unit_price=unit_price, order=order))

            try:
                with transaction.atomic():
                    if not duplicates:
                        Product.objects.bulk_create(new_products)

                        messages.success(request, 'You have created your order.')
                        return redirect('blog:order_detail', pk=order.pk)

                    else:
                        messages.warning(request, 'There are repeated products.')

                        context = {
                            'order_form': order_form,
                            'products_formset': product_formset
                        }

                        return render(request, 'blog/order_new.html', context)

            except IntegrityError:
                messages.error(request, 'There was an error saving your order.')

                context = {
                    'order_form': order_form,
                    'products_formset': product_formset
                }

                return render(request, 'blog/order_new.html', context)

    else:
        order_form = OrderForm(prefix="orderForm")
        products_formset = ProductFormSet(prefix="form")

    context = {
        'order_form': order_form,
        'products_formset': products_formset
    }

    return render(request, 'blog/order_new.html', context)


def setBordersCell(sheet):
    """
    Set borders cell function docstring.

    This function modify the cells of the Excel page to carry out an order and include the image of
    ICN2 logo.

    @param sheet: Excel sheet to modify the cells.

    @return: Excel sheet modified.
    """
    border_TopBottomThin = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    border_RightTopBottomThin = Border(right=Side(style='thin'), top=Side(style='thin'),
                                       bottom=Side(style='thin'))

    border_TopThin = Border(top=Side(style='thin'))

    border_BottomThin = Border(bottom=Side(style='thin'))

    border_TopThinBottomDouble = Border(top=Side(style='thin'), bottom=Side(style='double'))

    border_RightTopBottomMedium = Border(right=Side(style='medium'), top=Side(style='medium'),
                                         bottom=Side(style='medium'))

    sheet.cell('D6').border = border_TopBottomThin
    sheet.cell('E6').border = border_RightTopBottomThin
    sheet.cell('D7').border = border_TopBottomThin
    sheet.cell('E7').border = border_RightTopBottomThin
    sheet.cell('D8').border = border_TopBottomThin
    sheet.cell('E8').border = border_RightTopBottomThin
    sheet.cell('D9').border = border_TopBottomThin
    sheet.cell('E9').border = border_RightTopBottomThin
    sheet.cell('D10').border = border_TopBottomThin
    sheet.cell('E10').border = border_RightTopBottomThin
    sheet.cell('D11').border = border_TopBottomThin
    sheet.cell('E11').border = border_RightTopBottomThin
    sheet.cell('D12').border = border_TopBottomThin
    sheet.cell('E12').border = border_RightTopBottomThin

    sheet.cell('G12').border = border_RightTopBottomMedium

    sheet.cell('D14').border = border_TopBottomThin
    sheet.cell('E14').border = border_TopBottomThin
    sheet.cell('F14').border = border_TopBottomThin
    sheet.cell('G14').border = border_TopBottomThin

    sheet.cell('D18').border = border_TopBottomThin
    sheet.cell('E18').border = border_TopBottomThin
    sheet.cell('F18').border = border_TopBottomThin
    sheet.cell('G18').border = border_TopBottomThin
    sheet.cell('H18').border = border_RightTopBottomThin

    sheet.cell('G30').border = border_RightTopBottomThin

    sheet.cell('B36').border = border_TopThin
    sheet.cell('C36').border = border_TopThin
    sheet.cell('D36').border = border_TopThin
    sheet.cell('E36').border = border_TopThin

    sheet.cell('B37').border = border_BottomThin
    sheet.cell('C37').border = border_BottomThin
    sheet.cell('D37').border = border_BottomThin
    sheet.cell('E37').border = border_BottomThin
    sheet.cell('F37').border = border_BottomThin
    sheet.cell('G37').border = border_BottomThin

    for num in range(38, 65):
        numStr = str(num)
        sheet.cell('B' + numStr).border = border_TopBottomThin
        sheet.cell('C' + numStr).border = border_TopBottomThin
        sheet.cell('D' + numStr).border = border_TopBottomThin
        sheet.cell('E' + numStr).border = border_TopBottomThin
        sheet.cell('F' + numStr).border = border_TopBottomThin
        sheet.cell('G' + numStr).border = border_RightTopBottomThin

    sheet.cell('I67').border = border_TopBottomThin
    sheet.cell('I68').border = border_TopBottomThin
    sheet.cell('I69').border = border_TopThinBottomDouble

    img = Image('blog/static/images/icn2.png')
    # img.anchor(sheet.cell('H2'))
    sheet.add_image(img, 'H2')
    return sheet


@login_required
def order_new_next(request, pk):
    """
    order_new_next function docstring.

    This function shows the second step of the form to create a new order.

    @param request: HTML request page.

    @param pk: primary key of the new order.

    @return: First time, this shows the form to complete a new order. If the form is completed,
    the Excel sheet is completed with the data and return the details of the new order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    ProductFormSet = formset_factory(ProductForm, extra=order.number_product)
    if request.method == "POST":
        product = ProductForm()
        formset = ProductFormSet(request.POST)
        doc = openpyxl.load_workbook('blog/orderForm/Order_Form.xlsx')
        doc.get_sheet_names()
        sheet = doc.get_sheet_by_name('Order Form')
        sheet['C6'] = order.applicant
        sheet['C7'] = order.budget.name
        sheet['C10'] = order.type_of_purchase.name
        sheet['C12'] = order.payment_conditions.name
        sheet['C18'] = order.supplier.name
        num = 38
        nameFile = "FP_" + order.name
        if (formset.is_valid()):
            for form in formset:
                product = form.cleaned_data
                product = form.save(commit=False)
                numString = str(num)
                sheet['A' + numString] = product.description
                sheet['H' + numString] = product.quantity
                sheet['I' + numString] = product.unit_price
                num = num + 1
                product.order = order
                product.save()

            sheet = setBordersCell(sheet)
            doc.save('blog/orderForm/' + nameFile + '.xlsx')

            return redirect('blog:order_detail', pk=order.pk)
    else:
        formset = ProductFormSet()
    return render(request, 'blog/order_new_next.html', {'formset': formset})


@login_required
def order_edit(request, pk):
    """
    Order_edit function docstring.

    This function shows the form to modify an order.

    @param request: HTML request page.

    @param pk: primary key of the order to modify.

    @return: First time, this shows the form to edit the order information. If the form is
    completed, return the details of this order.

    @raise 404: order does not exists.
    """
    order_data = get_object_or_404(Order, pk=pk)

    ProductFormSet = formset_factory(ProductForm)

    products = Product.objects.filter(order=order_data.pk).order_by('created_date')
    product_data = [{'description': l.description, 'quantity': l.quantity,
                     'unit_price': l.unit_price, 'order': l.order} for l in products]

    if request.method == "POST":
        order_form = OrderForm(data=request.POST, instance=order_data, prefix="orderForm")
        product_formset = ProductFormSet(data=request.POST, prefix="form")

        if order_form.is_valid() and product_formset.is_valid():
            order = order_form.save(commit=False)
            order.author = request.user
            order.save()

            new_products = []
            duplicates = False

            for product_form in product_formset:
                description = product_form.cleaned_data.get('description')
                quantity = product_form.cleaned_data.get('quantity')
                unit_price = product_form.cleaned_data.get('unit_price')

                if description and quantity and unit_price:
                    for new_products_data in new_products:
                        if new_products_data.description == description:
                            duplicates = True

                    new_products.append(Product(description=description, quantity=quantity,
                                                unit_price=unit_price, order=order))

            try:
                with transaction.atomic():
                    if not duplicates:
                        Product.objects.filter(order=order).delete()
                        Product.objects.bulk_create(new_products)

                        messages.success(request, 'You have updated your order.')
                        return redirect('blog:order_detail', pk=order.pk)
                    else:
                        messages.warning(request, 'There are repeated products.')
                        return redirect('blog:order_edit', pk=order.pk)

            except IntegrityError:
                messages.error(request, 'There was an error saving your order.')
                return redirect('blog:order_detail', pk=order.pk)
    else:
        order_form = OrderForm(instance=order_data, prefix="orderForm")
        # products = Product.objects.filter(order=order_data.pk)
        noItem = False

        if not products.exists():
            noItem = True

        products_formset = ProductFormSet(initial=product_data, prefix="form")
        count = products.count()

    context = {
        'order_form': order_form,
        'products_formset': products_formset,
        'noItem': noItem,
        'count': count
    }

    return render(request, 'blog/order_edit.html', context)


@login_required
def order_send_email(request, pk):
    """
    Order_send_email function docstring.

    This function send an email with the order.

    @param request: HTML request page.

    @param pk: primary key of the order to send.

    @return: edit page of the order.

    @raise 404: order does not exists.
    """
    if request.method == "POST":
        order = get_object_or_404(Order, pk=pk)
        username = User.objects.get(username=order.author)
        sendEmail_form = SendEmailForm(data=request.POST)

        if sendEmail_form.is_valid():
            # sendEmail = sendEmail_form.save(commit=False)
            # print(sendEmail_form.cleaned_data.get('password'))
            fromaddr = username.email
            toaddrs = 'pexespada@gmail.com'
            subject = 'Formulario de pedido'
            message = "<p>Buenas,</p><p>Adjunto a este correo el formulario de pedido.</p>"
            msg = MIMEMultipart('related')
            msg['From'] = fromaddr
            msg['To'] = toaddrs
            msg['Subject'] = subject

            # Content-type:text/html
            message = MIMEText(message, 'html')
            msg.attach(message)
            # ADJUNTO
            file = 'blog/formulariosPedidos/FP_' + order.name + '.xlsx'
            if (os.path.isfile(file)):
                adjunto = MIMEBase('application', 'octet-stream')
                adjunto.set_payload(open(file, "rb").read())
                encode_base64(adjunto)
                adjunto.add_header('Content-Disposition',
                                   'attachment; filename = "%s"' % os.path.basename(file))
                msg.attach(adjunto)
            # ENVIAR
            server = smtplib.SMTP('mail.icn2.cat', 587)
            # protocolo de cifrado de datos utilizado por gmail
            server.starttls()
            # Credenciales
            # server.login(username.email, 'heriberto_20')
            server.login('icn2\\' + username.username, sendEmail_form.cleaned_data.get('password'))
            server.set_debuglevel(1)
            server.sendmail(fromaddr, toaddrs, msg.as_string())
            server.quit()
        return redirect('blog:order_detail', pk=pk)
    else:
        form = SendEmailForm()
        return render(request, 'blog/order_send_email.html', {'form': form})


@login_required
def order_remove(request, pk):
    """
    Order_remove function docstring.

    This function removes an order.

    @param request: HTML request page.

    @param pk: primary key of the order to remove.

    @return: list of orders.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk)
    order.delete()
    products.delete()
    return redirect('blog:order_list')


def run_list(request):
    """
    Run_list function docstring.

    This function shows the list of runs that are stored in this web app and they are ordered by
    creation date.

    @param request: HTML request page.

    @return: list of runs.
    """
    runs = Run.objects.filter(created_date__lte=timezone.now()).order_by('created_date').reverse()

    return render(request, 'blog/run_list.html', {'runs': runs})


def run_detail(request, pk):
    """
    Run_detail function docstring.

    This function shows the information of a run. BUT THIS FUNCTION IS NOT IN USE NOW.

    @param request: HTML request page.

    @param pk: primary key of the run.

    @return: one run.

    @raise 404: run does not exists.
    """
    run = get_object_or_404(Run, pk=pk)

    return render(request, 'blog/run_detail.html', {'run': run})


@login_required
def run_new(request):
    """
    Run_new function docstring.

    This function shows the form to create a new run. BUT THIS FUNCTION IS NOT IN USE NOW.

    @param request: HTML request page.

    @return: First time, this shows the form to a new run. If the form is completed, return
    the details of this new run.
    """
    if request.method == "POST":
        form = RunForm(request.POST)
        if form.is_valid():
            run = form.save(commit=False)
            run.save()
            return redirect('blog:run_detail', pk=run.pk)
    else:
        form = RunForm()
    return render(request, 'blog/run_edit.html', {'form': form})


@login_required
def run_edit(request, pk):
    """
    Run_edit function docstring.

    This function shows the form to modify a run. BUT THIS FUNCTION IS NOT IN USE NOW.

    @param request: HTML request page.

    @param pk: primary key of the run to modify.

    @return: First time, this shows the form to edit the run information. If the form is
    completed, return the details of this run.

    @raise 404: run does not exists.
    """
    run = get_object_or_404(Run, pk=pk)
    if request.method == "POST":
        form = RunForm(data=request.POST, instance=run)
        if form.is_valid():
            run = form.save(commit=False)
            run.save()
            return redirect('blog:run_detail', pk=run.pk)
    else:
        form = RunForm(instance=run)
    return render(request, 'blog/run_edit.html', {'form': form})


@login_required
def run_remove(request, pk):
    """
    Run_remove function docstring.

    This function removes a run. BUT THIS FUNCTION IS NOT IN USE NOW.

    @param request: HTML request page.

    @param pk: primary key of the run to remove.

    @return: list of runs.

    @raise 404: run does not exists.
    """
    run = get_object_or_404(Run, pk=pk)
    run.delete()
    return redirect('blog:run_list')


def run_chip_list(request, pk):
    """
    Run_chip_list function docstring.

    This function shows the list of chips of a run that are stored in this web app.

    @param request: HTML request page.

    @param pk: primary key of the run.

    @return: list of chips.
    """
    # chips = Chip.objects.filter(run=pk)
    chips = Chip.objects.filter(wafer=pk)

    return render(request, 'blog/chip_list.html', {'chips': chips})


def wafer_list(request, pk):
    """
    wafer_list function docstring.

    This function shows the list of wafers that are stored in this web app and they are ordered by
    creation date.

    @param request: HTML request page.

    @return: list of wafers.
    """
    wafers = Wafer.objects.filter(run=pk).order_by('created_date').reverse()

    return render(request, 'blog/wafer_list.html', {'wafers': wafers})


@login_required
def wafer_new(request):
    """
    wafer_new function docstring.

    This function shows the form to create a new wafer. This function uses 2 forms to create a new
    wafer (Run form and Wafer form), moreover, when the form is completed, this function
    search if the run or wafer already exists.

    @param request: HTML request page.

    @return: First time, this shows the form to a new wafer. If the form is completed and the run or
    wafer does not exists, return the list of chips of this new wafer.
    """
    if request.method == "POST":
        runForm = RunForm(request.POST, prefix='run')
        waferForm = WaferForm(request.POST, prefix='wafer')
        if runForm.is_valid():
            run = runForm.save(commit=False)
            # run_ex = Run.objects.get(run=run.run)
            run_ex = Run.objects.filter(run=run.run).exists()
            # print("Run: " + str(run_ex))
            if not run_ex:
                # print(run_ex)
                run.save()
            else:
                # run = run_ex
                run = Run.objects.get(run=run.run)
        if waferForm.is_valid():
            wafer = waferForm.save(commit=False)
            # wafer_ex = Wafer.objects.get(wafer=wafer.wafer, run=run)
            wafer_ex = Wafer.objects.filter(wafer=wafer.wafer, run=run).exists()
            # print("Wafer: " + str(wafer_ex))
            if not wafer_ex:
                # print(wafer_ex)
                wafer.run = run
                wafer.save()
                wafer_chip_new(run, wafer)
                return redirect('blog:wafer_chip_list', pk=wafer.pk)
            else:
                # wafer = wafer_ex
                wafer = Wafer.objects.get(wafer=wafer.wafer, run=run)
                return redirect('blog:wafer_detail_exist', pk=wafer.pk)
    else:
        runForm = RunForm(prefix='run')
        waferForm = WaferForm(prefix='wafer')
    return render(request, 'blog/wafer_edit.html', {'runForm': runForm, 'waferForm': waferForm})


def wafer_chip_new(run, wafer):
    """
    Wafer_chip_new function docstring.

    This function creates all chips of a wafer. This function uses a forms to create the new
    chips (Chip form).

    @param run: primary key of the run.
    @param wafer: primary key of the wafer.
    """
    chipName = ['CHIP1', 'CHIP2', 'CHIP3', 'CHIP4', 'CHIP5']
    chipGName = ['CHIPG1', 'CHIPG2', 'CHIPG3', 'CHIPG4', 'CHIPG5']
    chipPName = ['CHIP6P', 'CHIPG6P']

    for chip in chipName:
        chipForm = ChipForm()
        newChip = chipForm.save(commit=False)
        newChip.run = run
        newChip.wafer = wafer
        newChip.chip = chip
        # print(newChip)
        newChip.save()
        chip_waveguide_new(run, wafer, newChip)

    for chip in chipGName:
        chipForm = ChipForm()
        newChip = chipForm.save(commit=False)
        newChip.run = run
        newChip.wafer = wafer
        newChip.chip = chip
        # print(newChip)
        newChip.save()
        chip_waveguide_new(run, wafer, newChip)

    for chip in chipPName:
        chipForm = ChipForm()
        newChip = chipForm.save(commit=False)
        newChip.run = run
        newChip.wafer = wafer
        newChip.chip = chip
        # print(newChip)
        newChip.save()
        chip_waveguide_new(run, wafer, newChip)


def wafer_chip_list(request, pk):
    """
    Wafer_chip_list function docstring.

    This function shows the list of chips of a wafer that are stored in this web app.

    @param request: HTML request page.

    @param pk: primary key of the wafer.

    @return: list of chips.
    """
    # chips = Chip.objects.filter(run=pk)
    chips = Chip.objects.filter(wafer=pk)
    run = Wafer.objects.get(pk=pk)
    # print(run.run.pk)
    runback = True
    waferback = False

    return render(request, 'blog/chip_list.html', {'chips': chips, 'runPK': run.run.pk,
                                                   'runback': runback, 'waferback': waferback})


def chip_list(request):
    """
    Chip_list function docstring.

    This function shows the list of chips that are stored in this web app and they are ordered by
    creation date.

    @param request: HTML request page.

    @return: list of chips.
    """
    chips = Chip.objects.filter(created_date__lte=timezone.now()).order_by('created_date')

    return render(request, 'blog/chip_list.html', {'chips': chips})


def chip_detail(request, pk):
    """
    Chip_detail function docstring.

    This function shows the information of a chip.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @return: one chip.

    @raise 404: chip does not exists.
    """
    chip = get_object_or_404(Chip, pk=pk)
    runback = False
    waferback = True

    return render(request, 'blog/chip_detail.html', {'chip': chip, 'waferPK': chip.wafer.pk,
                                                     'runback': runback, 'waferback': waferback})


def chip_detail_exist(request, pk):
    """
    Chip_detail_exist function docstring.

    This function shows the information of a chip when a chip is created but this already exists.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @return: one chip.

    @raise 404: chip does not exists.
    """
    chip = get_object_or_404(Chip, pk=pk)

    return render(request, 'blog/chip_detail_exist.html', {'chip': chip})


@login_required
def chip_new(request):
    """
    Chip_new function docstring.

    This function shows the form to create a new chip. This function uses 3 forms to create a new
    chip (Run form, Wafer form and Chip form), moreover, when the form is completed, this function
    search if the run or wafer or chip already exists.

    @param request: HTML request page.

    @return: First time, this shows the form to a new chip. If the form is completed and the run or
    wafer or chip does not exists, return the details of this new chip.
    """
    if request.method == "POST":
        runForm = RunForm(request.POST, prefix='run')
        waferForm = WaferForm(request.POST, prefix='wafer')
        chipForm = ChipForm(request.POST, prefix='chip')
        # waveguideForm = WaveguideForm(request.POST, prefix='waveguide')
        if runForm.is_valid():
            run = runForm.save(commit=False)
            run_ex = Run.objects.get(run=run.run)
            print(run)
            if not run_ex:
                # print(run_ex)
                run.save()
            else:
                run = run_ex
        if waferForm.is_valid():
            wafer = waferForm.save(commit=False)
            wafer_ex = Wafer.objects.get(wafer=wafer.wafer, run=run)
            # print(wafer)
            if not wafer_ex:
                # print(wafer_ex)
                wafer.run = run
                wafer.save()
            else:
                wafer = wafer_ex
        if chipForm.is_valid():
            chip = chipForm.save(commit=False)
            chip_ex = Chip.objects.get(chip=chip.chip, wafer=wafer, run=run)
            # print(chip)
            if not chip_ex:
                # print(chip_ex)
                chip.run = run
                chip.wafer = wafer
                chip.save()
                return redirect('blog:chip_detail', pk=chip.pk)
            else:
                chip = chip_ex
                return redirect('blog:chip_detail_exist', pk=chip.pk)
    else:
        runForm = RunForm(prefix='run')
        waferForm = WaferForm(prefix='wafer')
        chipForm = ChipForm(prefix='chip')
        # waveguideFrom = WaveguideForm(prefix='waveguide')
    return render(request, 'blog/chip_edit.html', {'runForm': runForm, 'waferForm': waferForm,
                                                   'chipForm': chipForm})


@login_required
def chip_edit(request, pk):
    """
    Chip_edit function docstring.

    This function shows the form to modify a chip.

    @param request: HTML request page.

    @param pk: primary key of the chip to modify.

    @return: First time, this shows the form to edit the chip information. If the form is
    completed, return the details of this chip.

    @raise 404: chip does not exists.
    """
    chip = get_object_or_404(Chip, pk=pk)
    if request.method == "POST":
        form = ChipForm(data=request.POST, instance=chip)
        if form.is_valid():
            chip = form.save(commit=False)
            chip.save()
            return redirect('blog:chip_detail', pk=chip.pk)
    else:
        runForm = RunForm(instance=chip.run)
        waferForm = WaferForm(instance=chip.wafer)
        chipForm = ChipForm(instance=chip)
    return render(request, 'blog/chip_edit.html', {'runForm': runForm, 'waferForm': waferForm,
                                                   'chipForm': chipForm})


@login_required
def chip_remove(request, pk):
    """
    Chip_remove function docstring.

    This function removes a chip.

    @param request: HTML request page.

    @param pk: primary key of the chip to remove.

    @return: list of chips.

    @raise 404: chip does not exists.
    """
    chip = get_object_or_404(Run, pk=pk)
    chip.delete()
    return redirect('blog:chip_list')


def chip_waveguide_new(run, wafer, chip):
    """
    Chip_waveguide_new function docstring.

    This function creates all waveguides of a chip. This function uses a forms to create the new
    waveguides (Waveguide form).

    @param run: primary key of the run.
    @param wafer: primary key of the wafer.
    @param chip: primary key of the chip.
    """
    waveguideName = Name_Waveguide.objects.all()
    # print(waveguideName)

    for waveguide in waveguideName:
        # print(waveguide.name)
        waveguideForm = WaveguideForm()
        newWaveguide = waveguideForm.save(commit=False)
        newWaveguide.run = run
        newWaveguide.wafer = wafer
        newWaveguide.chip = chip
        newWaveguide.waveguide = waveguide
        newWaveguide.name = waveguide.name
        # print(newWaveguide)
        newWaveguide.save()


# def waveguide_list(request):
#
#     waveguides = Waveguide.objects.filter(
#        created_date__lte=timezone.now()).order_by('created_date')
#
#     return render(request, 'blog/waveguide_list.html', {'waveguides': waveguides})

def waveguide_list(request, pk):
    """
    Waveguide_list function docstring.

    This function shows the list of waveguides of a chip that are stored in this web app.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @return: list of waveguides of a chip.
    """
    waveguides = Waveguide.objects.filter(chip=pk)
    chipBack = True
    waveguideBack = False

    return render(request, 'blog/waveguide_list.html', {'waveguides': waveguides, 'chip': pk,
                                                        'chipBack': chipBack,
                                                        'waveguideBack': waveguideBack})


# def waveguide_detail(request, pk, pk2):
def waveguide_detail(request, pk):
    """
    Waveguide_detail function docstring.

    This function shows the information of a waveguide of a chip.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @param pk2: primary key of the waveguide.

    @return: one waveguide.

    @raise 404: waveguide does not exists.
    """
    waveguide = get_object_or_404(Waveguide, pk=pk)
    chipBack = False
    waveguideBack = True

    return render(request, 'blog/waveguide_detail.html', {'waveguide': waveguide,
                                                          'chipBack': chipBack,
                                                          'waveguideBack': waveguideBack})


def waveguide_detail_exist(request, pk, pk2):
    """
    Waveguide_detail_exist function docstring.

    This function shows the information of a waveguide of a chip when a waveguide is created but
    this already exists.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @param pk2: primary key of the waveguide.

    @return: one waveguide.

    @raise 404: waveguide does not exists.
    """
    waveguide = get_object_or_404(Waveguide, pk=pk2)

    return render(request, 'blog/waveguide_detail_exist.html', {'waveguide': waveguide})


@login_required
def waveguide_new(request, pk):
    """
    Waveguide_new function docstring.

    This function shows the form to create a new waveguide of a chip. When the form is completed,
    this function search if the waveguide already exists.

    @param request: HTML request page.

    @param pk: primary key of the chip.

    @return: First time, this shows the form to a new waveguide to a chip already created. If the
    form is completed and waveguide does not exists, return the details of this new waveguide.

    @raise 404: chip does not exists.
    """
    if request.method == "POST":
        # runForm = RunForm(request.POST, prefix='run')
        # waferForm = WaferForm(request.POST, prefix='wafer')
        waveguideForm = WaveguideForm(request.POST, prefix='waveguide')
        # waveguideForm = WaveguideForm(request.POST, prefix='waveguide')
        # if runForm.is_valid():
        #     run = runForm.save(commit=False)
        #     run_ex = Run.objects.get(run=run.run)
        #     print(run)
        #     if not run_ex:
        #         # print(run_ex)
        #         run.save()
        #     else:
        #         run = run_ex
        # if waferForm.is_valid():
        #     wafer = waferForm.save(commit=False)
        #     wafer_ex = Wafer.objects.get(wafer=wafer.wafer, run=run)
        #     # print(wafer)
        #     if not wafer_ex:
        #         # print(wafer_ex)
        #         wafer.run = run
        #         wafer.save()
        #     else:
        #         wafer = wafer_ex
        if waveguideForm.is_valid():
            waveguide = waveguideForm.save(commit=False)
            # print(waveguide)
            chip = Chip.objects.get(pk=pk)
            # print(chip.wafer)
            waveguide_ex = Waveguide.objects.get(waveguide=waveguide.waveguide, chip=chip,
                                                 wafer=chip.wafer, run=chip.run)
            # print(waveguide)
            if not waveguide_ex:
                # print(waveguide_ex)
                waveguide.run = chip.run
                waveguide.wafer = chip.wafer
                waveguide.chip = chip
                waveguide.name = waveguide.waveguide.name
                waveguide.save()
                return redirect('blog:waveguide_detail', pk=chip.pk, pk2=waveguide.pk)
            else:
                waveguide = waveguide_ex
                print(waveguide.pk)
                return redirect('blog:waveguide_detail_exist', pk=chip.pk, pk2=waveguide.pk)
    else:
        # runForm = RunForm(prefix='run')
        # waferForm = WaferForm(prefix='wafer')
        chip = get_object_or_404(Chip, pk=pk)
        waveguideForm = WaveguideForm(prefix='waveguide')
        # waveguideFrom = WaveguideForm(prefix='waveguide')
    return render(request, 'blog/waveguide_edit.html', {'waveguideForm': waveguideForm,
                                                        'chip': chip})


@login_required
def waveguide_edit(request, pk):
    """
    Waveguide_edit function docstring.

    This function shows the form to modify a waveguide.

    @param request: HTML request page.

    @param pk: primary key of the waveguide to modify.

    @return: First time, this shows the form to edit the waveguide information. If the form is
    completed, return the details of this waveguide.

    @raise 404: waveguide does not exists.
    """
    waveguide_data = Waveguide.objects.get(pk=pk)
    if request.method == "POST":
        form = WaveguideForm(data=request.POST)
        waveguide_data = Waveguide.objects.get(pk=pk)
        if form.is_valid():
            waveguide_data.amplitude = form.cleaned_data['amplitude']
            waveguide_data.offset = form.cleaned_data['offset']
            waveguide_data.frecuency = form.cleaned_data['frecuency']
            waveguide_data.i_up = form.cleaned_data['i_up']
            waveguide_data.i_down = form.cleaned_data['i_down']
            waveguide_data.slope = form.cleaned_data['slope']
            waveguide_data.visibility = form.cleaned_data['visibility']
            waveguide_data.noise = form.cleaned_data['noise']
            waveguide_data.lod = form.cleaned_data['lod']
            waveguide_data.save()
            return redirect('blog:waveguide_detail', pk=waveguide_data.pk)
    else:
        form = WaveguideForm(instance=waveguide_data)
    return render(request, 'blog/waveguide_edit.html', {'form': form})


@login_required
def waveguide_remove(request, pk):
    """
    Waveguide_remove function docstring.

    This function removes a waveguide.

    @param request: HTML request page.

    @param pk: primary key of the waveguide to remove.

    @return: list of waveguides.

    @raise 404: waveguide does not exists.
    """
    waveguide = get_object_or_404(Run, pk=pk)
    waveguide.delete()
    return redirect('blog:waveguide_list')
