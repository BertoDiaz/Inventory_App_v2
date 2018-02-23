"""
File name: order.py.

Name: Inventory App

Description: With this web application you can do the inventory of all
             the material of your laboratory or business. You can also
             place orders but this form is case-specific. Moreover,
             you can track all your manufacturing procedures such as
             wafer fabrication in this case.

Copyright (C) 2017  Heriberto J. Díaz Luis-Ravelo

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see http://www.gnu.org/licenses/.

Email: heriberto.diazluis@gmail.com
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from blog.views.search import get_query
from django.contrib.auth.models import User
from django.forms.formsets import formset_factory
from django.utils import timezone
from django.db import IntegrityError, transaction
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import smtplib
import os
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from blog.models import Order, Product, Budget, Type_of_purchase, Supplier
from blog.forms import OrderForm, ProductForm, SendEmailForm, UploadFileForm, SupplierForm


@login_required
def order_list(request):
    """
    Order_list function docstring.

    This function shows the list of orders carried out in this web app and they are ordered by
    creation date.

    @param request: HTML request page.

    @return: list of orders.
    """
    order_list = Order.objects.filter(created_date__lte=timezone.now()).order_by('created_date').reverse()

    # Show 25 contacts per page
    paginator = Paginator(order_list, 25)

    page = request.GET.get('page')
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        orders = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        orders = paginator.page(paginator.num_pages)

    return render(request, 'blog/order_list.html', {'orders': orders})


@login_required
def order_search(request):
    """
    Order_search function docstring.

    This function search the orders that are stored in this web app and they are
    ordered by name.

    @param request: HTML request page.

    @return: list of orders.
    """
    query_string = ''
    found_entries = None
    if ('searchfield' in request.GET) and request.GET['searchfield'].strip():
        query_string = request.GET['searchfield']
        try:
            query_string = User.objects.get(first_name=query_string)
            order_list = Order.objects.filter(author=query_string.pk).order_by('name')
        except ObjectDoesNotExist:
            try:
                query_string = Budget.objects.get(name=query_string)
                order_list = Order.objects.filter(budget=query_string.pk).order_by('name')
            except ObjectDoesNotExist:
                try:
                    query_string = Type_of_purchase.objects.get(name=query_string)
                    order_list = Order.objects.filter(type_of_purchase=query_string.pk).order_by('name')
                except ObjectDoesNotExist:
                    try:
                        query_string = Supplier.objects.get(name=query_string)
                        order_list = Order.objects.filter(supplier=query_string.pk).order_by('name')
                    except ObjectDoesNotExist:
                        entry_query = get_query(query_string, ['name', 'applicant'])
                        order_list = Order.objects.filter(entry_query).order_by('name')

    # Show 25 contacts per page
    paginator = Paginator(order_list, 25)

    page = request.GET.get('page')
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        orders = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        orders = paginator.page(paginator.num_pages)

    return render(request, 'blog/order_list.html', {'orders': orders})


@login_required
def order_detail(request, pk):
    """
    Order_detail function docstring.

    This function shows the information of an order.

    @param request: HTML request page.

    @param pk: primary key of the order.

    @return: one order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk).order_by('created_date')

    noItem = False
    if not products.exists():
        noItem = True

    backList = True
    backDetail = False

    return render(request, 'blog/order_detail.html', {'order': order, 'products': products,
                                                      'noItem': noItem, 'backList': backList,
                                                      'backDetail': backDetail})


@login_required
def order_new(request):
    """
    Order_new function docstring.

    This function shows the first step of the form to create a new order.

    @param request: HTML request page.

    @return: First time, this shows the form to a new order. If the form is completed, return
    the next step to finish the order.
    """
    ProductFormSet = formset_factory(ProductForm)

    if request.method == "POST":
        order_form = OrderForm(data=request.POST, prefix="orderForm")
        product_formset = ProductFormSet(data=request.POST, prefix="form")
        if order_form.is_valid() and product_formset.is_valid():
            order = order_form.save(commit=False)
            order.author = request.user

            supplier_form = SupplierForm(data=request.POST, prefix="supplierForm")

            if order.supplier.name == "SUPPLIER NOT REGISTERED" and not supplier_form.is_valid():

                products_formset = product_formset

                supplier_form = SupplierForm(prefix="supplierForm")
                addSupplier = True

                messages.warning(request, 'You have to write the next information about the supplier.')

                backList = True
                backDetail = False

                context = {
                    'order_form': order_form,
                    'products_formset': products_formset,
                    'supplier_form': supplier_form,
                    'backList': backList,
                    'backDetail': backDetail,
                    'addSupplier': addSupplier
                }

                return render(request, 'blog/order_new.html', context)

            else:

                if supplier_form.is_valid():
                    supplier = supplier_form.save(commit=False)
                    order.name_supplier = supplier.name
                    supplier.save()

                order.save()

                new_products = []
                duplicates = False

                doc = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/Order_Form.xlsx'))
                doc.get_sheet_names()
                sheet = doc.get_sheet_by_name('Order Form')
                sheet['C6'] = order.applicant
                sheet['C7'] = order.budget.name
                sheet['C10'] = order.type_of_purchase.name
                sheet['C12'] = order.payment_conditions.name
                sheet['C17'] = order.supplier.name

                if order.supplier.name == "SUPPLIER NOT REGISTERED":
                    sheet['C19'] = supplier.name
                    sheet['C20'] = supplier.attention
                    sheet['C21'] = supplier.address
                    sheet['C22'] = supplier.city_postCode
                    sheet['C23'] = supplier.phone
                    sheet['C24'] = supplier.fax
                    sheet['C25'] = supplier.email

                num = 37
                nameFile = "OF_" + order.name

                for product_form in product_formset:
                    description = product_form.cleaned_data.get('description')
                    quantity = product_form.cleaned_data.get('quantity')
                    unit_price = product_form.cleaned_data.get('unit_price')

                    if description and quantity and unit_price:
                        for new_products_data in new_products:
                            if new_products_data.description == description:
                                duplicates = True

                        numString = str(num)
                        sheet['A' + numString] = description
                        sheet['H' + numString] = quantity
                        sheet['I' + numString] = unit_price
                        num = num + 1

                        new_products.append(Product(description=description, quantity=quantity,
                                                    unit_price=unit_price, order=order))

                try:
                    with transaction.atomic():
                        if not duplicates:
                            Product.objects.bulk_create(new_products)

                            sheet = setBordersCell(sheet)

                            if not os.path.exists(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/')):
                                os.makedirs(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/'))

                            doc.save(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/' +
                                     nameFile + '.xlsx'))

                            messages.success(request, 'You have created your order.')
                            return redirect('blog:order_detail', pk=order.pk)

                        else:
                            messages.warning(request, 'There are repeated products.')

                            context = {
                                'order_form': order_form,
                                'products_formset': product_formset
                            }

                            return render(request, 'blog/order_new.html', context)

                except IntegrityError:
                    messages.error(request, 'There was an error saving your order.')

                    context = {
                        'order_form': order_form,
                        'products_formset': product_formset
                    }

                    return render(request, 'blog/order_new.html', context)

    else:
        order_form = OrderForm(prefix="orderForm")
        products_formset = ProductFormSet(prefix="form")

        backList = True
        backDetail = False
        addSupplier = False

    context = {
        'order_form': order_form,
        'products_formset': products_formset,
        'backList': backList,
        'backDetail': backDetail,
        'addSupplier': addSupplier
    }

    return render(request, 'blog/order_new.html', context)


def setBordersCell(sheet):
    """
    Set borders cell function docstring.

    This function modify the cells of the Excel page to carry out an order and include the image of
    ICN2 logo.

    @param sheet: Excel sheet to modify the cells.

    @return: Excel sheet modified.
    """
    border_TopBottomThin = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    border_RightTopBottomThin = Border(right=Side(style='thin'), top=Side(style='thin'),
                                       bottom=Side(style='thin'))

    border_TopThin = Border(top=Side(style='thin'))

    border_BottomThin = Border(bottom=Side(style='thin'))

    border_TopThinBottomDouble = Border(top=Side(style='thin'), bottom=Side(style='double'))

    border_RightTopBottomMedium = Border(right=Side(style='medium'), top=Side(style='medium'),
                                         bottom=Side(style='medium'))

    sheet.cell('D6').border = border_TopBottomThin
    sheet.cell('E6').border = border_RightTopBottomThin
    sheet.cell('D7').border = border_TopBottomThin
    sheet.cell('E7').border = border_RightTopBottomThin
    sheet.cell('D8').border = border_TopBottomThin
    sheet.cell('E8').border = border_RightTopBottomThin
    sheet.cell('D9').border = border_TopBottomThin
    sheet.cell('E9').border = border_RightTopBottomThin
    sheet.cell('D10').border = border_TopBottomThin
    sheet.cell('E10').border = border_RightTopBottomThin
    sheet.cell('D11').border = border_TopBottomThin
    sheet.cell('E11').border = border_RightTopBottomThin
    sheet.cell('D12').border = border_TopBottomThin
    sheet.cell('E12').border = border_RightTopBottomThin

    sheet.cell('G12').border = border_RightTopBottomMedium

    sheet.cell('D14').border = border_TopBottomThin
    sheet.cell('E14').border = border_TopBottomThin
    sheet.cell('F14').border = border_TopBottomThin
    sheet.cell('G14').border = border_RightTopBottomThin

    sheet.cell('D17').border = border_TopBottomThin
    sheet.cell('E17').border = border_TopBottomThin
    sheet.cell('F17').border = border_TopBottomThin
    sheet.cell('G17').border = border_TopBottomThin
    sheet.cell('H17').border = border_RightTopBottomThin

    sheet.cell('G29').border = border_RightTopBottomThin

    sheet.cell('B35').border = border_TopThin
    sheet.cell('C35').border = border_TopThin
    sheet.cell('D35').border = border_TopThin
    sheet.cell('E35').border = border_TopThin

    sheet.cell('B36').border = border_BottomThin
    sheet.cell('C36').border = border_BottomThin
    sheet.cell('D36').border = border_BottomThin
    sheet.cell('E36').border = border_BottomThin
    sheet.cell('F36').border = border_BottomThin
    sheet.cell('G36').border = border_BottomThin

    for num in range(37, 63):
        numStr = str(num)
        sheet.cell('B' + numStr).border = border_TopBottomThin
        sheet.cell('C' + numStr).border = border_TopBottomThin
        sheet.cell('D' + numStr).border = border_TopBottomThin
        sheet.cell('E' + numStr).border = border_TopBottomThin
        sheet.cell('F' + numStr).border = border_TopBottomThin
        sheet.cell('G' + numStr).border = border_RightTopBottomThin

    sheet.cell('I66').border = border_TopBottomThin
    sheet.cell('I67').border = border_TopBottomThin
    sheet.cell('I68').border = border_TopThinBottomDouble

    img = Image(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/images/icn2.png'))
    sheet.add_image(img, 'H2')
    return sheet


@login_required
def order_edit(request, pk):
    """
    Order_edit function docstring.

    This function shows the form to modify an order.

    @param request: HTML request page.

    @param pk: primary key of the order to modify.

    @return: First time, this shows the form to edit the order information. If the form is
    completed, return the details of this order.

    @raise 404: order does not exists.
    """
    order_data = get_object_or_404(Order, pk=pk)

    ProductFormSet = formset_factory(ProductForm)

    products = Product.objects.filter(order=order_data.pk).order_by('created_date')
    product_data = [{'description': l.description, 'quantity': l.quantity,
                     'unit_price': l.unit_price, 'order': l.order} for l in products]

    if request.method == "POST":
        order_form = OrderForm(data=request.POST, instance=order_data, prefix="orderForm")
        product_formset = ProductFormSet(data=request.POST, prefix="form")

        if order_form.is_valid() and product_formset.is_valid():
            order = order_form.save(commit=False)
            order.edited_by = request.user.username

            supplier_form = SupplierForm(data=request.POST, prefix="supplierForm")

            if order.supplier.name == "SUPPLIER NOT REGISTERED" and supplier_form.is_valid():

                products_formset = product_formset

                supplier_form = SupplierForm(prefix="supplierForm")
                addSupplier = True

                messages.warning(request, 'You have to write the next information about the supplier.')

                backList = True
                backDetail = False

                noItem = False

                if not products.exists():
                    noItem = True

                products_formset = ProductFormSet(initial=product_data, prefix="form")
                count = products.count()

                context = {
                    'order_form': order_form,
                    'products_formset': products_formset,
                    'supplier_form': supplier_form,
                    'backList': backList,
                    'backDetail': backDetail,
                    'addSupplier': addSupplier,
                    'noItem': noItem,
                    'count': count
                }

                return render(request, 'blog/order_edit.html', context)

            else:

                if supplier_form.is_valid():
                    supplier = supplier_form.save(commit=False)
                    order.name_supplier = supplier.name

                    supplier_data = Supplier.objects.filter(name=supplier.name)

                    if not supplier_data.exists():
                        supplier.save()

                order.save()

                new_products = []
                duplicates = False

                for product_form in product_formset:
                    description = product_form.cleaned_data.get('description')
                    quantity = product_form.cleaned_data.get('quantity')
                    unit_price = product_form.cleaned_data.get('unit_price')

                    if description and quantity and unit_price:
                        for new_products_data in new_products:
                            if new_products_data.description == description:
                                duplicates = True

                        new_products.append(Product(description=description, quantity=quantity,
                                                    unit_price=unit_price, order=order))

                try:
                    with transaction.atomic():
                        if not duplicates:
                            Product.objects.filter(order=order).delete()
                            Product.objects.bulk_create(new_products)

                            doc = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/OF_' + order.name + '.xlsx'))
                            doc.get_sheet_names()
                            sheet = doc.get_sheet_by_name('Order Form')
                            sheet['C6'] = order.applicant
                            sheet['C7'] = order.budget.name
                            sheet['C10'] = order.type_of_purchase.name
                            sheet['C12'] = order.payment_conditions.name
                            sheet['C17'] = order.supplier.name

                            if order.supplier.name == "SUPPLIER NOT REGISTERED":
                                sheet['C19'] = supplier.name
                                sheet['C20'] = supplier.attention
                                sheet['C21'] = supplier.address
                                sheet['C22'] = supplier.city_postCode
                                sheet['C23'] = supplier.phone
                                sheet['C24'] = supplier.fax
                                sheet['C25'] = supplier.email

                            num = 37

                            for product_form in product_formset:
                                description = product_form.cleaned_data.get('description')
                                quantity = product_form.cleaned_data.get('quantity')
                                unit_price = product_form.cleaned_data.get('unit_price')

                                if description and quantity and unit_price:
                                    for new_products_data in new_products:
                                        if new_products_data.description == description:
                                            duplicates = True

                                    numString = str(num)
                                    sheet['A' + numString] = description
                                    sheet['H' + numString] = quantity
                                    sheet['I' + numString] = unit_price
                                    num = num + 1

                            sheet = setBordersCell(sheet)

                            doc.save(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/OF_' +
                                     order.name + '.xlsx'))

                            messages.success(request, 'You have updated your order.')
                            return redirect('blog:order_detail', pk=order.pk)
                        else:
                            messages.warning(request, 'There are repeated products.')
                            return redirect('blog:order_edit', pk=order.pk)

                except IntegrityError:
                    messages.error(request, 'There was an error saving your order.')
                    return redirect('blog:order_detail', pk=order.pk)
    else:
        order_form = OrderForm(instance=order_data, prefix="orderForm")

        if order_data.supplier.name == "SUPPLIER NOT REGISTERED":
            supplier_data = Supplier.objects.get(name=order_data.name_supplier)
            supplier_form = SupplierForm(instance=supplier_data, prefix="supplierForm")

        noItem = False

        if not products.exists():
            noItem = True

        products_formset = ProductFormSet(initial=product_data, prefix="form")
        count = products.count()

        backList = False
        backDetail = True

    if order_data.supplier.name == "SUPPLIER NOT REGISTERED":
        context = {
            'order': order_data,
            'order_form': order_form,
            'products_formset': products_formset,
            'supplier_form': supplier_form,
            'noItem': noItem,
            'count': count,
            'backList': backList,
            'backDetail': backDetail
        }
    else:
        context = {
            'order': order_data,
            'order_form': order_form,
            'products_formset': products_formset,
            'noItem': noItem,
            'count': count,
            'backList': backList,
            'backDetail': backDetail
        }

    return render(request, 'blog/order_edit.html', context)


@login_required
def order_send_email(request, pk):
    """
    Order_send_email function docstring.

    This function send an email with the order.

    @param request: HTML request page.

    @param pk: primary key of the order to send.

    @return: edit page of the order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    username = User.objects.get(username=order.author)

    if request.method == "POST":
        sendEmail_form = SendEmailForm(data=request.POST)

        if sendEmail_form.is_valid():
            fromaddr = username.email
            toaddrs = 'pexespada@gmail.com'
            subject = 'Order form'

            if (order.file_exists):
                message_1 = "<p>Dear Jessica,</p><p>" + request.POST['message'] + "</p>"
                message_2 = "<p></p><p>Best regards,</p><p>" + username.first_name + "</p>"
                message = message_1 + message_2
            else:
                message = "<p>Dear Jessica,</p><p>" + request.POST['message'] + "</p><p>Best regards,</p><p>" + username.first_name + "</p>"

            msg = MIMEMultipart('related')
            msg['From'] = fromaddr
            msg['To'] = toaddrs
            msg['Subject'] = subject

            # Content-type:text/html
            message = MIMEText(message, 'html')
            msg.attach(message)

            # ADJUNTO
            orderForm = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/OF_' + order.name + '.xlsx')
            if (os.path.isfile(orderForm)):
                adjunto = MIMEBase('application', 'octet-stream')
                adjunto.set_payload(open(orderForm, "rb").read())
                encode_base64(adjunto)
                adjunto.add_header('Content-Disposition',
                                   'attachment; filename = "%s"' % os.path.basename(orderForm))
                msg.attach(adjunto)

            if order.file_exists:
                uploadFile = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static/orderForm/' + order.author.username + '/' + order.name_file_attach)
                if (os.path.isfile(uploadFile)):
                    adjunto = MIMEBase('application', 'octet-stream')
                    adjunto.set_payload(open(uploadFile, "rb").read())
                    encode_base64(adjunto)
                    adjunto.add_header('Content-Disposition',
                                       'attachment; filename = "%s"' % os.path.basename(uploadFile))
                    msg.attach(adjunto)

            try:
                # ENVIAR
                server = smtplib.SMTP('mail.icn2.cat', 587)
                # protocolo de cifrado de datos utilizado por gmail
                server.starttls()
                # Credenciales
                server.login('icn2\\' + username.username, sendEmail_form.cleaned_data.get('password'))
                server.set_debuglevel(1)
                server.sendmail(fromaddr, toaddrs, msg.as_string())
                server.quit()

                order.order_sent = True
                order.save()

                messages.success(request, 'The order has been sent successfully.')

            except:
                messages.error(request, 'Username or password is incorrect.')

        return redirect('blog:order_detail', pk=pk)
    else:
        form = SendEmailForm()

        backList = False
        backDetail = True

        if (order.file_exists):
            message_1 = "Dear Jessica,"
            message_2 = "Attached to this email the order form and budget."
            message_3 = "Best regards,"
            message_4 = str(username.first_name)

        else:
            message_1 = "Dear Jessica,"
            message_2 = "Attached to this email the order form."
            message_3 = "Best regards,"
            message_4 = str(username.first_name)

        if (order.order_sent):
            messages.warning(request, 'This order has already been sent.')

        context ={
        'form': form,
        'order': order,
        'backList': backList,
        'backDetail': backDetail,
        'message_1': message_1,
        'message_2': message_2,
        'message_3': message_3,
        'message_4': message_4
        }

        return render(request, 'blog/order_send_email.html', context)


@login_required
def order_notify(request, pk):
    """
    Order_notify function docstring.

    This function notify to an group that an order is going to be send.

    @param request: HTML request page.

    @param pk: primary key of the order to notify.

    @return: detail page of the order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    username = User.objects.get(username=order.author)
    users = User.objects.exclude(username=order.author)

    usersAdd = []

    for userID in users:
        usersAdd.append(userID.email)

    if request.method == "POST":
        sendEmail_form = SendEmailForm(data=request.POST)

        if sendEmail_form.is_valid():
            fromaddr = username.email
            toaddrs = usersAdd
            subject = 'New Order'

            message = "<p>Hello everybody,</p><p>" + request.POST['message'] + "</p><p>Best regards,</p><p>" + username.first_name + "</p>"

            msg = MIMEMultipart('related')
            msg['From'] = fromaddr
            msg['To'] = ", ".join(toaddrs)
            msg['Subject'] = subject

            # Content-type:text/html
            message = MIMEText(message, 'html')
            msg.attach(message)

            try:
                # ENVIAR
                server = smtplib.SMTP('mail.icn2.cat', 587)
                # protocolo de cifrado de datos utilizado por gmail
                server.starttls()
                # Credenciales
                server.login('icn2\\' + username.username, sendEmail_form.cleaned_data.get('password'))
                server.set_debuglevel(1)
                server.sendmail(fromaddr, toaddrs, msg.as_string())
                server.quit()

                order.group_notified = True
                order.save()

                messages.success(request, 'The group has been notified successfully.')

            except:
                messages.error(request, 'Username or password is incorrect.')

        return redirect('blog:order_detail', pk=pk)
    else:
        form = SendEmailForm()

        backList = False
        backDetail = True

        message_1 = "Hello everybody,"
        message_2 = "I am about to place an order for " + str(order.supplier) + "."
        message_3 = " Please, let me know if anybody need anything."
        message_4 = "Best regards,"
        message_5 = str(username.first_name)

        if (order.group_notified):
            messages.warning(request, 'The group has already been notified.')

        context ={
        'form': form,
        'order': order,
        'backList': backList,
        'backDetail': backDetail,
        'message_1': message_1,
        'message_2': message_2,
        'message_3': message_3,
        'message_4': message_4,
        'message_5': message_5
        }

        return render(request, 'blog/order_notify_email.html', context)


@login_required
def order_add_file(request, pk):
    """
    Order_add_file function docstring.

    This function add a file to the order.

    @param request: HTML request page.

    @param pk: primary key of the order.

    @return: detail page of the order.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)

    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            handle_uploaded_file(request.FILES['file'], order)

            order.file_exists = True
            f = request.FILES['file']
            name, extension = os.path.splitext(f.name)
            order.name_file_attach = 'UF_' + order.name + extension
            order.save()

            messages.success(request, 'The file have been upload to your order.')

        return redirect('blog:order_detail', pk=pk)

    else:
        form = UploadFileForm()

        backList = False
        backDetail = True

        return render(request, 'blog/order_add_file.html', {'form': form, 'order': order,
                                                            'backList': backList,
                                                            'backDetail': backDetail})


def handle_uploaded_file(f, order):
    """
    Handle_uploaded_file function docstring.

    This function save a file in the server.

    @param f: file to upload.

    @param order: order of the file.
    """
    name, extension = os.path.splitext(f.name)
    with open(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'orderForm/' + order.author.username + '/UF_' +
              order.name + extension), 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


@login_required
def order_remove(request, pk):
    """
    Order_remove function docstring.

    This function removes an order.

    @param request: HTML request page.

    @param pk: primary key of the order to remove.

    @return: list of orders.

    @raise 404: order does not exists.
    """
    order = get_object_or_404(Order, pk=pk)
    products = Product.objects.filter(order=order.pk)
    order.delete()
    products.delete()
    return redirect('blog:order_list')
